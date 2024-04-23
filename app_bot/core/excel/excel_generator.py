import io
import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tortoise.fields.relational import BackwardFKRelation
from openpyxl.utils import get_column_letter
from core.database.models import Report


async def create_excel(model):
    entries = await model.all()

    file_in_memory = io.BytesIO()
    book = Workbook()
    sheet = book.active

    # get model headers
    headers = []
    for field in model._meta.fields_map.values():
        if type(field) != BackwardFKRelation:
            headers.append(field.model_field_name)
    sheet.append(headers)

    # add users data
    for entry in entries:
        row = []
        for field_name in headers:
            cell = getattr(entry, field_name)
            if type(cell) == datetime.datetime:
                cell: datetime.datetime = cell.replace(tzinfo=None)
            row.append(cell)

        sheet.append(row)

    book.save(file_in_memory)
    file_in_memory.seek(0)

    return file_in_memory


async def create_excel_after_checking(reports: list[Report]):
    file_in_memory = io.BytesIO()
    book = Workbook()
    sheet = book.active

    sheet.append(['Айди отчета', 'Экспонат', 'Статус', 'Комментарий', 'ФИО проверяющего'])

    # add data
    is_empty = True
    for report in reports:
        if report.status != Report.StatusType.work:
            is_empty = False
            sheet.append(
                [
                    report.id,
                    (await report.exhibit).name,
                    report.status.value,
                    report.description,
                    (await report.creator).fio
                ]
            )

    book.save(file_in_memory)
    file_in_memory.seek(0)

    return file_in_memory, is_empty


async def sort_reports_by_date(reports):
    report_data = {}
    for report in reports:
        key = (report.created_at.date(), (await report.exhibit).name)
        report_data[key] = report.status.value


    # fill dates and statuses for the month
    all_exhibits = sorted(set(name for _, name in report_data.keys()))
    start_date = min(date for date, _ in report_data.keys())
    end_date = max(date for date, _ in report_data.keys())
    current_date = start_date

    while current_date <= end_date:
        for exhibit_name in all_exhibits:
            key = (current_date, exhibit_name)
            if key not in report_data:
                prev_day = current_date - datetime.timedelta(days=1)
                prev_key = (prev_day, exhibit_name)
                report_data[key] = report_data.get(prev_key, '')
        current_date += datetime.timedelta(days=1)

    return report_data


# /stats reports excel
async def create_main_reports_excel(reports: list[Report]):
    data_dict = await sort_reports_by_date(reports=reports)

    file_in_memory = io.BytesIO()
    book = Workbook()
    ws: Worksheet = book.active

    # fill the first row with dates
    dates = sorted(set(date for date, _ in data_dict.keys()))
    ws.append([''] + [str(date) for date in dates])

    # fill cells (date, exhibit__name)
    exhibit_names = sorted(set(name for _, name in data_dict.keys()))
    for name in exhibit_names:
        row = [name]
        for date in dates:
            status = data_dict.get((date, name), '')
            row.append(status)
        ws.append(row)

    # set columns row width
    for column in range(ws.min_column, ws.max_column):
        ws.column_dimensions[get_column_letter(column)].width = 15

    book.save(file_in_memory)
    file_in_memory.seek(0)

    return file_in_memory
